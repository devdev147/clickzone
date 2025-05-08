#pip install pandas
#pip install xlrd
#pip install xlwt
#pip install openpyxl
import pandas as pd #to export import df to and from excel files
import openpyxl #to handle excel files directly from python without using dataframe

#use pandas to read excel file into a dataframe or write dataframe to an excel file
#use ExcelWriter object to write dataframe into multiple worksheets of a workbook or to append to an exiting excel file
#use openpyxl to append to an existing excel file

#df = pd.read_excel("sample.xlsx")
#df = pd.read_excel(r'C:\Users\Ron\Desktop\products.xlsx') 
#df1.to_excel("output.xlsx")  
#df1.to_excel("output.xlsx",sheet_name='Sheet_name_1')  

def readingexcel(xlsxfilename):
    #reading excel file and its properties directly using openpyxl without importing it to a dataframe
    wb = openpyxl.load_workbook(xlsxfilename) #Loading Workbook
    ws1 = wb.active #Openning active worksheet (first sheet in workbook)
    ws2 = wb['Sheet2'] #Openning a specific worksheet
    rows = ws1.max_row #Counting number of rows in a worksheet
    cols = ws1.max_column #Counting number of columns in a worksheet
    print('Rows: ',rows,' Columns: ',cols)
    val = ws1['A1'].value #Reading data from a cell
    print('Value in cell A1: ',val)
    rowvals = [ws1.cell(row=1,column=i).value for i in range(2,cols+1)] #Reading data from multiple cells - all columns of a row
    print('Values of all columns in row 1: ',rowvals)
    colvals = [ws1.cell(row=i,column=2).value for i in range(2,rows+1)] #Reading multiple rows of a specific column
    print('Values of all rows of column 2: ',colvals)
    #Reading data from multiple rows and columns (specifying a range)
    rangevals = list()
    for cellval in ws1.iter_rows(min_row=2, max_row=rows, min_col=2, max_col=cols, values_only=True):
        rangevals.append(cellval)
    print("Values in a range B2:C3 - ",rangevals)
    for v1,v2,v3 in rangevals:
        print("{:<8}{:<20}{:<10}".format(v1,v2,v3))

def writingexcel(xlsxfilename):
    #Writing to a cell in an excel file with Openpyxl
    wb = openpyxl.load_workbook(xlsxfilename) 
    ws = wb['Sheet1'] 
    ws['F1'] = 'Sum of Sales'
    #Creating (appending) a new column
    ws.cell(row=1, column=6, value = 'TOTAL')
    c = 2
    for r in range(2,ws.max_row+1):
        total = ((ws.cell(row=r, column=c).value)+
                 (ws.cell(row=r, column=c+1).value)+
                 (ws.cell(row=r, column=c+2).value))
        ws.cell(row=r,column=6).value=total
    ws.cell(row=1, column=7, value = 'TOTAL-1')
    for r in range(2,ws.max_row+1):
        total = 0
        for c in range(2,ws.max_column+1):
            if ws.cell(row=r, column=c).value is not None:
                total += ws.cell(row=r, column=c).value
        ws.cell(row=r,column=7).value=total
    #Creating (appending) a new row
    newrow = (4,15,26,39)
    ws.append(newrow)
    #Printing the last row in the workbook
    values = [ws.cell(row=ws.max_row,column=i).value for i in range(1,ws.max_column+1)]
    print(values)
    #Deleting rows
    ws.delete_rows(ws.max_row-1, 1) # row number, number of rows to delete
    
    wb.save('sample.xlsx')

def applyingformula(xlsxfilename):
    wb = openpyxl.load_workbook(xlsxfilename) 
    ws = wb['Sheet1'] 
    #Creating excel formulas with Openpyxl
    ws['A6'] = 'Average'
    ws['B6'] = '= AVERAGE(B2:B4)'
    #
    ws['A7'] = 'Count Not Null'
    ws['B7'] = '=COUNTA(B2:B4)'
    #
    ws['A8'] = 'Count score<15'
    ws['B8']= '=COUNTIF(B2:B4, "<15")'
    wb.save('sample.xlsx')

def getsetworksheetproperties(xlsxfilename):
    wb = openpyxl.load_workbook(xlsxfilename) 
    ws = wb['Sheet1'] 
    #print the name of the active sheet
    print(ws.title)
    #rename the name of worksheet
    ws.title ='New Worksheet Title'
    #list all the worksheets of a workbook as an array
    print(wb.sheetnames)
    #create an empty worksheet
    wb.create_sheet('SheetNew')
    #delet a worksheet
    wb.remove(wb['Sheet4'])
    wb.save('sample.xlsx')

def formattingcells(xlsxfilename):
    #Formatting Cells
    #Changing Font Sizes and Styles
    from openpyxl.styles import Font
    from openpyxl.styles import colors
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side
    from openpyxl.formatting.rule import CellIsRule
    wb = openpyxl.load_workbook(xlsxfilename) 
    ws = wb['Sheet1']            
    ws['A1'].font = Font(bold=True, size=12)
    #creating a for loop to iterate through all the columns in the first row
    for cell in ws["1:1"]: 
        cell.font = Font(bold=True, size=12)
    #change font colors in Openpyxl using hex codes
    ws['A1'].font = Font(color = 'FF0000',bold=True, size=12) ## red
    ws['A2'].font = Font(color = '0000FF') ## blue
    # changing background color of a cell
    ws["A1"].fill = PatternFill('solid', start_color="38e3ff") # light blue background color
    #adding cell border 
    my_border = Side(border_style="thin", color="000000")
    ws["A1"].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
    #Conditional Formatting
    fill = PatternFill(start_color='90EE90',end_color='90EE90',fill_type='solid') # specify background color
    ws.conditional_formatting.add('G2:K16594', CellIsRule(operator='greaterThan', formula=[8], fill=fill)) # include formatting rule
    wb.save('sample.xlsx')

def createchart(xlsxfilename):
    #Adding charts to an Excel file
    from openpyxl.chart import Reference
    from openpyxl.chart import BarChart
    wb = openpyxl.load_workbook(xlsxfilename) 
    ws = wb['Sheet1'] 
    #specify the data values (Y-axis values) for plotting
    values = Reference(ws,     # worksheet object   
                   min_col=2,  # minimum column where your values begin
                   max_col=2,  # maximum column where your values end
                   min_row=2,  # minimum row you’d like to plot from
                   max_row=4)  # maximum row you’d like to plot from
    #set parameters for the chart categories (X-axis labels)
    cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=4)
    #create bar chart object by including data values and categories 
    chart = BarChart()
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    #set chart titles and set the location to create it in the Excel sheet
    # set the title of the chart
    chart.title = "ENG Marks"
    # set the title of the x-axis
    chart.x_axis.title = "Student"
    # set the title of the y-axis
    chart.y_axis.title = "Eng Marks"
    # the top-left corner of the chart is anchored to cell K2 
    ws.add_chart(chart,"K2")
    
    wb.save("sample.xlsx")

def protectworksheet(xlsxfilename):
    from openpyxl.styles import Protection
    wb = openpyxl.load_workbook(xlsxfilename) 
    ws = wb['Sheet1']
    #ws = wb.active
    #Lock the entire worksheet
    ws.protection.sheet = True
    #Unlock the specific cells that can be edited
    #cell.protection = Protection(locked=False)
    #After locking the sheet, iterate over the cells to be unlocked
    #the entire sheet will be locked but the specified columns will remain editable   
    for col in ['B','C']:
        for cell in ws[col]:
            cell.protection = Protection(locked=False)
    #entire worksheet protection
    ws = wb["Sheet2"]
    ws.protection.sheet = True
    ws.protection.enable()
    #ws.protection.disable()
    #setting protection password
    ws.protection.password = '...'
    wb.save('sample.xlsx')

def df2excel(df1):
    df1.to_excel("output.xlsx", index=True, header=True, sheet_name='Sheet1')  

def multipledf2multipleexcelsheets(df1,df2):
    # create ExcelWriter object to write to multiple sheets in a workbook
    with pd.ExcelWriter('output.xlsx') as writer:  
        df1.to_excel(writer, sheet_name='Sheet1')
        df2.to_excel(writer, sheet_name='Sheet2')

def appenddf2existingexcel(df1,df2):
    # append data to an existing Excel file by using ExcelWriter object with 'a' mode
    with pd.ExcelWriter("output.xlsx", engine="openpyxl", mode="a") as writer:
        #df.to_excel(writer, sheet_name="name", startrow=num, startcol=num)
        df1.to_excel(writer, sheet_name='Sheet3')
        df2.to_excel(writer, sheet_name='Sheet4')

def excel2df(xlsfilename):
    df = pd.read_excel( xlsfilename,
                        sheet_name='Sheet1',
                        index_col=0,
                        dtype={"Name":str,"Roll":int}
                        ) #usecols=['Roll','Name'],
    print(df)
# data -------------------------------------------
df1 = pd.DataFrame([[11, 21, 31],
                    [12, 22, 32],
                    [31, 32, 33]],
                    index=['1','2','3'],
                    columns=['A','B','C']
                    )
df2 = pd.DataFrame([['x11', 'x21', 'x31'],
                    ['x12', 'x22', 'x32'],
                    ['x31', 'x32', 'x33']],
                    index=['x1','x2','x3'],
                    columns=['xA','xB','xC']
                    )    
# function calls ---------------------------------
excel2df("data.xlsx")
df2excel(df1)   
multipledf2multipleexcelsheets(df1,df2)
appenddf2existingexcel(df1,df2)
readingexcel("output.xlsx")
writingexcel("output.xlsx")
applyingformula("output.xlsx")
getsetworksheetproperties("output.xlsx")
formattingcells("output.xlsx")
createchart("output.xlsx")
protectworksheet("output.xlsx")
