#pip install pandas
#pip install xlrd
#pip install openpyxl
import pandas as pd 
df = pd.read_excel("sample.xlsx")
print(df)
# Here 0th column will be extracted
df = pd.read_excel("sample.xlsx",index_col = 0)  
print(df)
#set dataypes of columns
df = pd.read_excel('sample.xlsx',dtype = {"Products": str,"Price":float})
print(df)
df = pd.read_excel(r'Path of Excel file\File name.xlsx', sheet_name='your Excel sheet name')
print(df)
data = pd.read_excel(r'C:\Users\Ron\Desktop\products.xlsx') 
df = pd.DataFrame(data, columns=['product_name'])
print(df)

#---------------------------------
import openpyxl 
#Loading the Workbook
wb = openpyxl.load_workbook('videogamesales.xlsx')
#Openning the active worksheet (the first sheet in the workbook)
ws = wb.active
#Openning a specific worksheet
ws = wb['vgsales']
#Count the number of rows and columns in the worksheet
print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
#Reading Data From a Cell
print('The value in cell A1 is: '+ws['A1'].value)
#Reading Data From Multiple Cells
values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)
#Printing multiple rows of a specific column
data=[ws.cell(row=i,column=2).value for i in range(2,12)]
print(data)
#printing first ten rows in a range of columns in the spreadsheet
#Reading data from a range of cells (from column 1 to 6)

my_list = list()

for value in ws.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6, 
    values_only=True):
    my_list.append(value)
    
for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
    (print ("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1,ele2,ele3,ele4,ele5,ele6)))

'''
This code reads data from a range of cells in a worksheet using the iter_rows() method from the openpyxl library.
•  The iter_rows() method takes in the parameters min_row, max_row, min_col, and max_col to specify the range of cells to read from.
•  The values_only=True parameter ensures that only the values of the cells are returned, not any formatting or other metadata.
•  The values are then appended to an empty list called my_list.
•  Finally, the values in my_list are printed out using a for loop that iterates through each element in the list.
•  The print() function uses string formatting to align the values in columns of a fixed width.
•  The < symbol is used to left-align the values within their respective columns, and the {:8} syntax specifies the width of each column.
•  Overall, this code reads data from a range of cells in a worksheet and prints it out in a formatted table.
'''
            
#Writing to Excel Files with Openpyxl
#Writing to a Cell
ws['K1'] = 'Sum of Sales'
#specify row and column position of the cell to write to
ws.cell(row=1, column=11, value = 'Sum of Sales')
#save changes in the worksheet
wb.save('videogamesales.xlsx')
#Creating a New Column
row_position = 2
col_position = 7
total_sales = ((ws.cell(row=row_position, column=col_position).value)+
               (ws.cell(row=row_position, column=col_position+1).value)+
               (ws.cell(row=row_position, column=col_position+2).value)+
               (ws.cell(row=row_position, column=col_position+3).value))
ws.cell(row=2,column=11).value=total_sales
wb.save('videogamesales.xlsx')
#creating a for loop to sum the sales values in every row
row_position = 1
for i in range(1, ws.max_row):
    row_position += 1
    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value
    total_sales = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)
    ws.cell(row=row_position, column=11).value = total_sales
wb.save("videogamesales.xlsx")

#Appending New Rows
#To append a new row to the workbook, create a tuple with the values and write it to the sheet
new_row = (1,'The Legend of Zelda',1986,'Action','Nintendo',3.74,0.93,1.69,0.14,6.51,6.5)
ws.append(new_row)
wb.save('videogamesales.xlsx')
#confirming the appending of data by printing the last row in the workbook
values = [ws.cell(row=ws.max_row,column=i).value for i in range(1,ws.max_column+1)]
print(values)
#Deleting Rows
ws.delete_rows(ws.max_row, 1) # row number, number of rows to delete
wb.save('videogamesales.xlsx')

#Creating Excel Formulas with Openpyxl
#create a new column called “Average Sales” to calculate the average total video game sales in all markets
ws['P1'] = 'Average Sales'
ws['P2'] = '= AVERAGE(K2:K16220)'
wb.save('videogamesales.xlsx')
#“COUNTA” function in Excel counts cells that are populated within a specific range. Let’s use it to find the number of records between E2 and E16220:
ws['Q1'] = "Number of Populated Cells" 
ws['Q2'] = '=COUNTA(E2:E16220)'
wb.save('videogamesales.xlsx')
#COUNTIF is a commonly used Excel function that counts the number of cells that meet a specific condition. Let’s use it to count the number of games in this dataset with the “Sports” genre:
ws['R1'] = 'Number of Rows with Sports Genre'
ws['R2'] = '=COUNTIF(E2:E16220, "Sports")'
wb.save('videogamesales.xlsx')           
#SUMIF to find the total “Sum of Sales” generated by sports games using the SUMIF function:
ws['S1'] = 'Total Sports Sales'
ws['S2'] = '=SUMIF(E2:E16220, "Sports",K2:K16220)'
wb.save('videogamesales.xlsx')
#CEILING function in Excel rounds a number up to the nearest specified multiple. Let’s round up the total amount of sales generated by sports games using this function:
ws['T1'] = 'Rounded Sum of Sports Sales'
ws['T2'] = '=CEILING(S2,25)'
wb.save('videogamesales.xlsx')

#Working with Sheets in Openpyxl
#print the name of the active sheet we are currently working with using Openpyxl’s title attribute:
print(ws.title)
#rename this worksheet using the following lines of code:
ws.title ='Video Game Sales Data'
wb.save('videogamesales.xlsx')
# list all the worksheets in the workbook (gives an array listing the names of all the worksheets in the file)
print(wb.sheetnames)
#create a new empty worksheet
wb.create_sheet('Empty Sheet') # create an empty worksheet named 'Empty Sheet'
print(wb.sheetnames) # print sheet names again
wb.save('videogamesales.xlsx')
#Deleting a Worksheet
wb.remove(wb['Empty Sheet'])
print(wb.sheetnames)
wb.save('videogamesales.xlsx')
#Duplicating a Worksheet (creating a copy of an existing worksheet)



#Adding Charts to an Excel File
from openpyxl.chart import Reference
from openpyxl.chart import BarChart
#access the required worksheet
ws = wb['Total Sales by Genre'] 
#specify the data values (Y-axis values) for plotting
values = Reference(ws,         # worksheet object   
                   min_col=2,  # minimum column where your values begin
                   max_col=2,  # maximum column where your values end
                   min_row=1,  # minimum row you’d like to plot from
                   max_row=13) # maximum row you’d like to plot from
#set parameters for the chart categories (X-axis labels)
cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=13)
#create bar chart object by including data values and categories 
chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.set_categories(cats)
#set chart titles and set the location to create it in the Excel sheet
# set the title of the chart
chart.title = "Total Sales"
# set the title of the x-axis
chart.x_axis.title = "Genre"
# set the title of the y-axis
chart.y_axis.title = "Total Sales by Genre"
# the top-left corner of the chart is anchored to cell F2 
ws.add_chart(chart,"D2")
# save the file 
wb.save("videogamesales.xlsx")



#Formatting Cells
#Changing Font Sizes and Styles
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import CellIsRule
            
ws = wb['Video Game Sales Data']
ws['A1'].font = Font(bold=True, size=12)
wb.save('videogamesales.xlsx')
#creating a for loop to iterate through all the columns in the first row
for cell in ws["1:1"]: 
    cell.font = Font(bold=True, size=12)
wb.save('videogamesales.xlsx')
#change font colors in Openpyxl using hex codes
ws['A1'].font = Font(color = 'FF0000',bold=True, size=12) ## red
ws['A2'].font = Font(color = '0000FF') ## blue
wb.save('videogamesales.xlsx')
change the background color of a cell, you can use Openpyxl’s PatternFill module:
# changing background color of a cell
ws["A1"].fill = PatternFill('solid', start_color="38e3ff") # light blue background color
wb.save('videogamesales.xlsx')            
#adding cell border 
my_border = Side(border_style="thin", color="000000")
ws["A1"].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
wb.save("videogamesales.xlsx")
#Conditional Formatting
fill = PatternFill(start_color='90EE90',end_color='90EE90',fill_type='solid') # specify background color
ws.conditional_formatting.add('G2:K16594', CellIsRule(operator='greaterThan', formula=[8], fill=fill)) # include formatting rule
wb.save('videogamesales.xlsx')

            
#Lock the entire sheet:
ws.protection.sheet = True
#unlock the cells that can be edited
#cell.protection = Protection(locked=False)
#After locking the sheet, iterate over the cells to be unlocked
#the entire sheet will be locked but the specified columns will remain editable
ws = wb["RFI"]
ws.protection.sheet = True
for col in ['U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
        for cell in ws[col]:
            cell.protection = Protection(locked=False)
#worksheet protection
ws = wb.active
ws.protection.sheet = True
ws.protection.enable()
ws.protection.disable()
#setting the protection password
#setting password using 'openpxyl.worksheet.protection.SheetProtection.password()' property
ws = wb.active
ws.protection.password = '...'

#
wb=openpyxl.Workbook()
sheet1=wb.active
sheet1.cell(3,3).value = "Though shalt not overwrite"
sheet1.protection.sheet = True
sheet1.protection.password = 'test'
wb.save('Book1.xlsx')

